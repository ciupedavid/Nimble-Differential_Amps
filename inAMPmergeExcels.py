import time
import csv
import openpyxl
import json
import pandas as pd
import os

with open(r'inAMPNoise.json') as d:
    paths = json.load(d)['Nimble'][0]

path_file = pd.read_csv(
    paths['project_location'] + '\\' + paths['device'] + '\\' + 'Amplifier - Output Referred Noise.csv')
path_file.to_excel(paths['project_location'] + '\\' + paths['device'] + '\\' + 'Amplifier - Output Referred Noise.xlsx',
                   index=None, header=True)

excel_path = paths['project_location'] + '\\' + paths['device'] + '\\' + 'Amplifier - Output Referred Noise.xlsx'

file = openpyxl.load_workbook(excel_path)
sheet_obj = file.active
sheet_obj.delete_cols(3)
time.sleep(0.25)
sheet_obj.delete_cols(3)
time.sleep(0.25)
sheet_obj.delete_cols(3)
time.sleep(0.25)
sheet_obj.delete_cols(3)
file.save(excel_path)

text_file = paths['project_location'] + '\\' + paths['device'] + '\\' + 'Noise_Simulation.txt'
output_file = paths['project_location'] + '\\' + paths['device'] + '\\' + 'Noise_Simulation.xlsx'

wb = openpyxl.Workbook()
ws = wb.worksheets[0]

with open(text_file, 'r') as data:
    reader = csv.reader(data, delimiter='\t')
    for row in reader:
        ws.append(row)

wb.save(output_file)

path1 = output_file
path2 = excel_path

wb1 = openpyxl.load_workbook(filename=path1)
ws1 = wb1.worksheets[0]

wb2 = openpyxl.load_workbook(filename=path2)
ws2 = wb2.create_sheet(ws1.title)

for row in ws1:
    for cell in row:
        ws2[cell.coordinate].value = cell.value

wb2.save(path2)

xl = openpyxl.load_workbook(excel_path)
sheet1 = xl['Sheet']
sheet2 = xl['Sheet1']

columnA = []

for i in range(1, 1000, 1):
    columnA.append(sheet1.cell(row=i, column=1).value)

for i in range(1, 1000, 1):

    for i in range(1, 1000, 1):
        sheet2.cell(row=i, column=4).value = columnA[i - 1]

columnB = []

for i in range(1, 1000, 1):
    columnB.append(sheet1.cell(row=i, column=2).value)

for i in range(1, 1000, 1):

    for i in range(1, 1000, 1):
        sheet2.cell(row=i, column=5).value = columnB[i - 1]

xl.save(excel_path)

xr = pd.read_excel(excel_path)

xr.rename(columns={xr.columns[0]: 'Frequency (Hz)'}, inplace=True)
xr.rename(columns={xr.columns[1]: 'Total (nV/rt(Hz))'}, inplace=True)
xr.rename(columns={xr.columns[2]: ''}, inplace=True)
xr.rename(columns={xr.columns[3]: 'Ltspice Freq'}, inplace=True)
xr.rename(columns={xr.columns[4]: 'Ltspice V(onoise)'}, inplace=True)
# rename_column.drop(columns={rename_column.columns[0]}, axis=1, inplace=True)

xr['Total (nV/rt(Hz))'] = xr['Total (nV/rt(Hz))'].apply(lambda x: x * 1e9)
xr['Ltspice V(onoise)'] = xr['Ltspice V(onoise)'].apply(lambda x: x * 1e9)

xr.to_excel(excel_path, sheet_name='G ' + paths['gain'])

time.sleep(3)

extra_files_remove = paths['project_location'] + paths['device']
zip_remove_nimble = extra_files_remove + '\\' + 'Nimble - ' + paths['device'] + ' G' + paths['gain'] + '.zip'
zip_remove_ltspice = extra_files_remove + '\\' + 'LTspice - ' + paths['device'] + ' G' + paths['gain'] + '.zip'
os.remove(extra_files_remove + '\\' + 'AC_Simulation.asc')
os.remove(extra_files_remove + '\\' + 'Noise_Simulation.asc')
os.remove(extra_files_remove + '\\' + 'Transient_Simulation.asc')
os.remove(extra_files_remove + '\\' + 'Noise_Simulation.log')
os.remove(extra_files_remove + '\\' + 'Noise_Simulation.raw')
os.remove(extra_files_remove + '\\' + 'Noise_Simulation.txt')
os.remove(extra_files_remove + '\\' + 'Noise_Simulation.xlsx')
os.remove(extra_files_remove + '\\' + 'Noise_Simulation.op.raw')
os.remove(extra_files_remove + '\\' + 'Amplifier - Output Referred Noise.csv')
os.rmdir(extra_files_remove + '\\' + 'Raw Data')
os.remove(zip_remove_nimble)
os.remove(zip_remove_ltspice)
