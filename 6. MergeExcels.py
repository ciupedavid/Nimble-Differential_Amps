import openpyxl as xl
import json
# This script copies data from two EXCELS into one excel
# Data from AC_Simulation will be displayed on sheet1
# Data from Transfer Function will be displayed on sheet11

with open(r'paths.json') as d:
    paths = json.load(d)['Unzip'][0]

path1 = paths['ac_simulationexceledit']
path2 = paths['transfer_function']

wb1 = xl.load_workbook(filename=path1)
ws1 = wb1.worksheets[0]

wb2 = xl.load_workbook(filename=path2)
ws2 = wb2.create_sheet(ws1.title)

for row in ws1:
    for cell in row:
        ws2[cell.coordinate].value = cell.value

wb2.save(path2)